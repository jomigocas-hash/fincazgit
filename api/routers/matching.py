"""
Motor de Matching: compara criterios de demanda contra la oferta disponible.
Devuelve inmuebles ordenados por score (0-100).
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from db import get_conn

router = APIRouter()


# --- Modelos ---

class DemandSearch(BaseModel):
    # Obligatorios
    city:           str
    operation_type: str
    property_type:  str
    max_price:      float
    min_bedrooms:   int
    min_area_m2:    float
    # Opcionales
    stratum:        Optional[int]   = None
    locality:       Optional[str]   = None
    neighborhood:   Optional[str]   = None
    parking:        Optional[int]   = None
    covered_garage: Optional[bool]  = None
    # Datos del solicitante (para guardar)
    role:           str             = "buyer"
    client_name:    Optional[str]   = None
    client_email:   Optional[str]   = None
    agent_id:       Optional[int]   = None


# --- Lógica de scoring ---

def compute_score(prop: dict, demand: DemandSearch) -> int:
    """
    Calcula un score de 0-100 entre un inmueble y los criterios de demanda.
    """
    score = 0

    # Convertir Decimal a float para evitar errores de tipo
    price    = float(prop["price"])    if prop["price"]    else None
    area     = float(prop["area_m2"])  if prop["area_m2"]  else None
    bedrooms = int(prop["bedrooms"])   if prop["bedrooms"] is not None else None
    stratum  = int(prop["stratum"])    if prop["stratum"]  is not None else None
    parking  = int(prop["parking"])    if prop["parking"]  is not None else None

    # Precio (30 pts)
    if price and demand.max_price:
        if price <= demand.max_price:
            score += int(30 * (price / demand.max_price))
        else:
            over = (price - demand.max_price) / demand.max_price
            score += max(0, int(30 * (1 - over * 2)))

    # Área (20 pts)
    if area and demand.min_area_m2:
        if area >= demand.min_area_m2:
            score += 20
        else:
            score += int(20 * (area / demand.min_area_m2))

    # Habitaciones (20 pts)
    if bedrooms is not None and demand.min_bedrooms:
        if bedrooms >= demand.min_bedrooms:
            score += 20
        else:
            score += int(20 * (bedrooms / demand.min_bedrooms))

    # Barrio (10 pts)
    if demand.neighborhood and prop["neighborhood"]:
        if demand.neighborhood.lower() in prop["neighborhood"].lower():
            score += 10

    # Localidad (10 pts)
    if demand.locality and prop["neighborhood"]:
        if demand.locality.lower() in prop["neighborhood"].lower():
            score += 10

    # Estrato (5 pts)
    if demand.stratum and stratum:
        if stratum == demand.stratum:
            score += 5

    # Parqueadero (5 pts)
    if demand.parking and parking:
        if parking >= demand.parking:
            score += 5

    return min(score, 100)


# --- Endpoints ---

@router.post("/search")
def match_search(demand: DemandSearch, limit: int = Query(20, le=100)):
    """
    Recibe criterios de demanda y devuelve los inmuebles con mejor score.
    """
    sql = """
        SELECT id, canonical_id, source_portal, property_type, operation_type,
               city, neighborhood, address,
               CASE WHEN location IS NOT NULL THEN ST_Y(location::geometry) ELSE NULL END AS latitude,
               CASE WHEN location IS NOT NULL THEN ST_X(location::geometry) ELSE NULL END AS longitude,
               area_m2, bedrooms, bathrooms, parking, floor, stratum,
               price, price_per_m2, currency, admin_fee,
               title, url, image_urls, published_at
        FROM properties
        WHERE city = %(city)s
          AND operation_type = %(operation_type)s
          AND property_type  = %(property_type)s
          AND is_active = TRUE
          AND price IS NOT NULL
          AND price <= %(max_price)s
          AND (%(min_area_m2)s IS NULL OR area_m2 >= %(min_area_m2)s)
          AND (%(min_bedrooms)s IS NULL OR bedrooms >= %(min_bedrooms)s)
          AND (%(locality)s IS NULL OR neighborhood ILIKE %(locality_like)s)
          AND (%(neighborhood)s IS NULL OR neighborhood ILIKE %(neighborhood_like)s)
          AND (%(stratum)s IS NULL OR stratum = %(stratum)s)
        ORDER BY price ASC
        LIMIT 500
    """
    params = {
        "city":              demand.city.lower(),
        "operation_type":    demand.operation_type.lower(),
        "property_type":     demand.property_type.lower(),
        "max_price":         demand.max_price,
        "min_area_m2":       demand.min_area_m2,
        "min_bedrooms":      demand.min_bedrooms,
        "locality":          demand.locality or None,
        "locality_like":     f"%{demand.locality}%" if demand.locality else None,
        "neighborhood":      demand.neighborhood or None,
        "neighborhood_like": f"%{demand.neighborhood}%" if demand.neighborhood else None,
        "stratum":           demand.stratum or None,
    }

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            candidates = [dict(r) for r in cur.fetchall()]

    # Calcular score para cada candidato
    scored = []
    for prop in candidates:
        s = compute_score(prop, demand)
        scored.append({**prop, "match_score": s})

    # Ordenar por score descendente
    scored.sort(key=lambda x: x["match_score"], reverse=True)

    return {
        "total_candidates": len(scored),
        "results": scored[:limit],
    }


@router.post("/save")
def save_search(demand: DemandSearch):
    """
    Guarda una búsqueda de demanda (para agentes o compradores registrados).
    """
    sql = """
        INSERT INTO demand_searches (
            role, agent_id, client_name, client_email,
            city, operation_type, property_type,
            max_price, min_bedrooms, min_area_m2,
            stratum, locality, neighborhood, parking, covered_garage
        ) VALUES (
            %(role)s, %(agent_id)s, %(client_name)s, %(client_email)s,
            %(city)s, %(operation_type)s, %(property_type)s,
            %(max_price)s, %(min_bedrooms)s, %(min_area_m2)s,
            %(stratum)s, %(locality)s, %(neighborhood)s,
            %(parking)s, %(covered_garage)s
        ) RETURNING id
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, demand.model_dump())
            search_id = cur.fetchone()["id"]
        conn.commit()

    return {"id": search_id, "message": "Búsqueda guardada"}


@router.get("/searches")
def list_searches(agent_id: Optional[int] = Query(None), limit: int = 50):
    """Lista las búsquedas guardadas, opcionalmente filtradas por agente."""
    sql = """
        SELECT * FROM demand_searches
        WHERE is_active = TRUE
        {where}
        ORDER BY created_at DESC
        LIMIT %(limit)s
    """
    params = {"limit": limit}
    where = ""
    if agent_id:
        where = "AND agent_id = %(agent_id)s"
        params["agent_id"] = agent_id

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql.format(where=where), params)
            rows = cur.fetchall()

    return [dict(r) for r in rows]


def _build_excel(results: list, demand: DemandSearch) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Matching"

    BLUE_DARK   = PatternFill("solid", fgColor="1E3A5F")
    BLUE_MID    = PatternFill("solid", fgColor="2563EB")
    GREEN_FILL  = PatternFill("solid", fgColor="16A34A")
    BLUE_FILL   = PatternFill("solid", fgColor="2563EB")
    YELLOW_FILL = PatternFill("solid", fgColor="CA8A04")
    GRAY_FILL   = PatternFill("solid", fgColor="6B7280")
    ROW_ALT     = PatternFill("solid", fgColor="F1F5F9")
    WHITE_FONT  = Font(color="FFFFFF", bold=True)

    ws.merge_cells("A1:K1")
    ws["A1"] = "REPORTE DE MATCHING — INMUEBLES COLOMBIA"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = BLUE_DARK
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Ciudad: {demand.city.title()}  |  Operación: {demand.operation_type.title()}  |  Tipo: {demand.property_type.title()}"
    ws["A2"].font = Font(color="FFFFFF", size=10)
    ws["A2"].fill = BLUE_MID
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:K3")
    criterios = f"Precio máx ${demand.max_price:,.0f}  |  Hab. mín {demand.min_bedrooms}  |  Área mín {demand.min_area_m2} m²"
    if demand.locality:    criterios += f"  |  Localidad: {demand.locality}"
    if demand.neighborhood: criterios += f"  |  Barrio: {demand.neighborhood}"
    if demand.stratum:     criterios += f"  |  Estrato: {demand.stratum}"
    ws["A3"] = criterios
    ws["A3"].font = Font(color="FFFFFF", size=9)
    ws["A3"].fill = PatternFill("solid", fgColor="334155")
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    ws.append([])

    headers = ["#", "Score", "ID", "Canonical ID", "Título", "Barrio", "Precio", "Precio/m²", "Área m²", "Hab.", "Baños", "Estrato", "Portal"]
    ws.append(headers)
    hr = ws.max_row
    for col in range(1, 14):
        c = ws.cell(row=hr, column=col)
        c.fill = BLUE_DARK
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center")

    for i, prop in enumerate(results, 1):
        score  = prop["match_score"]
        precio = float(prop["price"])       if prop["price"]       else 0
        ppm2   = float(prop["price_per_m2"]) if prop["price_per_m2"] else 0
        ws.append([
            i, f"{score}%", prop["id"], prop.get("canonical_id") or "—",
            prop["title"], prop["neighborhood"] or "—",
            f"${precio:,.0f}", f"${ppm2:,.0f}", prop["area_m2"] or "—",
            prop["bedrooms"] or "—", prop["bathrooms"] or "—",
            prop["stratum"] or "—", prop["source_portal"],
        ])
        r = ws.max_row
        fill = ROW_ALT if i % 2 == 0 else None
        score_fill = GREEN_FILL if score >= 80 else BLUE_FILL if score >= 60 else YELLOW_FILL if score >= 40 else GRAY_FILL
        for col in range(1, 14):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(horizontal="center" if col in [1,2,3,9,10,11,12] else "left")
            if fill: cell.fill = fill
        ws.cell(row=r, column=2).fill = score_fill
        ws.cell(row=r, column=2).font = Font(color="FFFFFF", bold=True)

    for i, w in enumerate([5, 8, 7, 34, 40, 22, 18, 14, 10, 7, 7, 9, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append([])
    total_props = len(results)
    avg = sum(float(p["price"]) for p in results if p["price"]) / max(total_props, 1)
    ws.append(["", f"Total: {total_props}", "", "", f"Precio promedio: ${avg:,.0f}"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/export/excel")
def export_excel(demand: DemandSearch, limit: int = Query(100, le=500)):
    """Ejecuta el matching y descarga los resultados en Excel."""
    result  = match_search(demand, limit=limit)
    results = result["results"]
    excel   = _build_excel(results, demand)
    filename = f"matching_{demand.city}_{demand.operation_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
